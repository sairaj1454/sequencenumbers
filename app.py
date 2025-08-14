import os
from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_marketing_group(wers_code):
    """Determine marketing group based on WERS code prefix"""
    if not wers_code or not isinstance(wers_code, str):
        return ""
        
    wers_prefix = wers_code.split('#')[0] if '#' in wers_code else wers_code
    
    mapping = {
        '#T#': 'fp-overflow',
        'YZU': 'fp-vehicle.bodycode',
        'YZA': 'fp-vehicle.pepcode',
        'YCW': 'fp-powertrain.rearaxleratio',
        'ITS': 'fp-interior.material',
        'YCM': 'fp-interior.material',
        'TR-': 'fp-powertrain.transmission',
        'SW1': 'fp-exterior.wheels',
        'ST1': 'fp-exterior.tire',
        'SE#': 'fp-vehicle.series',
        'PAA': 'fp-exterior.paint',
        'Entity': 'fp-entity',
        'EN-': 'fp-powertrain.engine',
        'DR-': 'fp-powertrain.wheeldrive',
        '000': 'fp-interior.color'
    }
    
    # Check for exact matches first
    for prefix, group in mapping.items():
        if wers_code.startswith(prefix):
            return group
    
    # Check for partial matches
    for prefix, group in mapping.items():
        if prefix in wers_code:
            return group
    
    return ""

def process_excel(file_path):
    # Load the workbook and select the active worksheet
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    # Find header row and column indices
    header_row = None
    headers = {}
    
    # First, find the header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
        for cell in row:
            if cell.value == 'Top Family WERS Code':
                header_row = row_idx
                break
        if header_row:
            break
    
    if not header_row:
        raise ValueError("Could not find 'Top Family WERS Code' in the Excel file")
    
    # Now get all headers and their column indices
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[cell.value] = col_idx
    
    # Check if required columns exist
    required_columns = ['Top Family WERS Code', 'Sequence Number', 'Marketing Group']
    for col in required_columns:
        if col not in headers:
            raise ValueError(f"Required column '{col}' not found in the Excel file")
    
    # Check for at least one of the description columns
    desc_columns = ['Short Description_CA-EN', 'Short Description']
    has_desc_column = any(col in headers for col in desc_columns)
    if not has_desc_column:
        raise ValueError("Could not find either 'Short Description_CA-EN' or 'Short Description' column in the Excel file")
    
    # Get the description column to use (prioritize 'Short Description_CA-EN' if both exist)
    desc_column = next((col for col in desc_columns if col in headers), None)
    
    # Dictionary to track sequence numbers for each WERS code
    sequence_numbers = {}
    
    # First pass: Collect all WERS codes that have descriptions
    wers_codes_with_descriptions = set()
    for row in ws.iter_rows(min_row=header_row + 1):
        wers_cell = row[headers['Top Family WERS Code'] - 1]
        desc_cell = row[headers[desc_column] - 1] if desc_column else None
        
        if wers_cell.value and desc_cell and desc_cell.value and str(desc_cell.value).strip() != '':
            wers_codes_with_descriptions.add(wers_cell.value)
    
    # Second pass: Update sequence numbers and marketing groups
    for row in ws.iter_rows(min_row=header_row + 1):
        wers_cell = row[headers['Top Family WERS Code'] - 1]
        desc_cell = row[headers[desc_column] - 1] if desc_column else None
        seq_cell = row[headers['Sequence Number'] - 1]
        
        if wers_cell.value and desc_cell and desc_cell.value and str(desc_cell.value).strip() != '':
            wers_code = wers_cell.value
            
            # If this is a new WERS code, initialize its sequence number
            if wers_code not in sequence_numbers:
                sequence_numbers[wers_code] = 100  # Start from 100 for each WERS code
            
            # Set the sequence number and increment for next time
            seq_cell.value = sequence_numbers[wers_code]
            sequence_numbers[wers_code] += 5  # Increment by 5 for next entry
            
            # Set marketing group based on WERS code
            if 'Marketing Group' in headers:
                marketing_group_cell = row[headers['Marketing Group'] - 1]
                marketing_group_cell.value = get_marketing_group(str(wers_code))
    
    return wb

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            return redirect(request.url)
            
        file = request.files['file']
        
        # If user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            return redirect(request.url)
            
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Process the file
            try:
                processed_wb = process_excel(filepath)
                output_filename = 'processed_' + filename
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                
                # Save the workbook
                processed_wb.save(output_path)
                
                return render_template('index.html', 
                                    download_file=output_filename,
                                    show_download=True)
            except Exception as e:
                error_message = f"<div style='color: red; padding: 20px;'>"
                error_message += f"<h2>Error processing file:</h2>"
                error_message += f"<p>{str(e)}</p>"
                error_message += "<p>Please make sure your Excel file has the required columns: 'Top Family WERS Code', 'Authored', and 'Sequence Number'</p>"
                error_message += "</div>"
                return error_message
    
    return render_template('index.html', show_download=False)

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(
        os.path.join(app.config['UPLOAD_FOLDER'], filename),
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
